"""Monta o workbook Excel final a partir das tabelas extraidas de cada PDF.

Estrutura funcional herdada do modelo (modelo/Setor21_20260720.xlsx): titulo,
subtotais por grupo, TOTAL GERAL, freeze panes, autofiltro e bloco de
contabilizacao dinamica. Camada visual "executivo azul-marinho": banner navy,
somente separadores horizontais (sem grade vertical), gridlines ocultas e
tabelas com cabecalhos distintos em abas proprias.
"""

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from ..normalize import cell_value, is_currency_column
from .dynamic_array import inject_dynamic_arrays

# --- Design tokens base (paleta vibrante esmeralda & acentos radiantes) ---
FONT_NAME = "Calibri"
NAVY = "FF047857"          # banner, cabecalho da tabela, TOTAL GERAL (Verde Esmeralda Vivo)
NAVY_DARK = "FF064E3B"     # borda inferior do cabecalho (Verde Escuro)
NAVY_MID = "FF0D9488"      # acentos secundarios (Teal Vivo)
INK = "FF0F172A"           # texto principal dos dados (Slate 900)
MUTED = "FF475569"         # subtitulos e textos secundarios (Slate 600)
HAIRLINE = "FFE2E8F0"      # separadores horizontais (Slate 200)
ZEBRA = "FFECFDF5"         # faixa alternada (zebra) viva e suave (Emerald 50)
GROUP_FILL = "FFD1FAE5"    # linha "Total" de grupo (Emerald 100)
SUMMARY_FILL = "FFF0FDF4"  # corpo do card do painel dinamico (Emerald 50)
KPI_FILL = "FFF0FDF4"      # fundo dos cartoes de KPI (Emerald 50)
CARD_BORDER = "FF6EE7B7"   # borda dos cartoes / painel (Emerald 300)
BAR_COLOR = "FF10B981"     # cor das barras de dados (Emerald 500)
BANNER_SUBTITLE = "FFE2E8F0"
KPI_ACCENT = "FF10B981"
WHITE = "FFFFFFFF"
TAB_COLOR = "047857"
TAB_COLOR_WARN = "EF4444"
TAB_PALETTE = ["047857", "2563EB", "D97706", "7C3AED", "0891B2", "DC2626"]

# --- Paleta rotativa de cores por GRUPO (dentro de uma mesma aba) ---
# Cada entrada: {fill: fundo do subtotal, font: cor do texto, zebra: faixa alternada, accent: borda lateral}
GROUP_COLORS: list[dict[str, str]] = [
    {"fill": "FFD1FAE5", "font": "FF047857", "zebra": "FFECFDF5", "accent": "FF10B981"},  # Esmeralda
    {"fill": "FFDBEAFE", "font": "FF1D4ED8", "zebra": "FFEFF6FF", "accent": "FF3B82F6"},  # Azul
    {"fill": "FFFDE68A", "font": "FF92400E", "zebra": "FFFFFBEB", "accent": "FFF59E0B"},  # Âmbar
    {"fill": "FFEDE9FE", "font": "FF6D28D9", "zebra": "FFF5F3FF", "accent": "FFA78BFA"},  # Violeta
    {"fill": "FFCFFAFE", "font": "FF0E7490", "zebra": "FFECFEFF", "accent": "FF22D3EE"},  # Ciano
    {"fill": "FFFECDD3", "font": "FFB91C1C", "zebra": "FFFFF1F2", "accent": "FFF87171"},  # Coral
    {"fill": "FFFED7AA", "font": "FF9A3412", "zebra": "FFFFF7ED", "accent": "FFFB923C"},  # Laranja
    {"fill": "FFE0E7FF", "font": "FF3730A3", "zebra": "FFEEF2FF", "accent": "FF818CF8"},  # Índigo
]

# --- Paletas tematicas por cor de aba ---
# Cada entrada: (banner_dark, banner, accent, zebra, group_fill, kpi_fill, card_border, bar_color)
TAB_THEMES: dict[str, dict] = {
    # Verde Esmeralda
    "047857": {
        "banner":     "FF047857",
        "banner_dk":  "FF064E3B",
        "accent":     "FF0D9488",
        "zebra":      "FFECFDF5",
        "group_fill": "FFD1FAE5",
        "kpi_fill":   "FFF0FDF4",
        "card_bdr":   "FF6EE7B7",
        "bar_color":  "10B981",
        "kpi_text":   "FF047857",
    },
    # Azul Royal
    "2563EB": {
        "banner":     "FF2563EB",
        "banner_dk":  "FF1E3A8A",
        "accent":     "FF3B82F6",
        "zebra":      "FFEFF6FF",
        "group_fill": "FFDBEAFE",
        "kpi_fill":   "FFEFF6FF",
        "card_bdr":   "FF93C5FD",
        "bar_color":  "3B82F6",
        "kpi_text":   "FF1D4ED8",
    },
    # Âmbar / Dourado
    "D97706": {
        "banner":     "FFD97706",
        "banner_dk":  "FF92400E",
        "accent":     "FFF59E0B",
        "zebra":      "FFFFFBEB",
        "group_fill": "FFFDE68A",
        "kpi_fill":   "FFFFFBEB",
        "card_bdr":   "FFFCD34D",
        "bar_color":  "F59E0B",
        "kpi_text":   "FFB45309",
    },
    # Violeta
    "7C3AED": {
        "banner":     "FF7C3AED",
        "banner_dk":  "FF4C1D95",
        "accent":     "FFA78BFA",
        "zebra":      "FFF5F3FF",
        "group_fill": "FFEDE9FE",
        "kpi_fill":   "FFF5F3FF",
        "card_bdr":   "FFC4B5FD",
        "bar_color":  "A78BFA",
        "kpi_text":   "FF6D28D9",
    },
    # Ciano / Azul-piscina
    "0891B2": {
        "banner":     "FF0891B2",
        "banner_dk":  "FF164E63",
        "accent":     "FF22D3EE",
        "zebra":      "FFECFEFF",
        "group_fill": "FFCFFAFE",
        "kpi_fill":   "FFECFEFF",
        "card_bdr":   "FF67E8F9",
        "bar_color":  "22D3EE",
        "kpi_text":   "FF0E7490",
    },
    # Vermelho Coral
    "DC2626": {
        "banner":     "FFDC2626",
        "banner_dk":  "FF7F1D1D",
        "accent":     "FFF87171",
        "zebra":      "FFFFF1F2",
        "group_fill": "FFFECDD3",
        "kpi_fill":   "FFFFF1F2",
        "card_bdr":   "FFFCA5A5",
        "bar_color":  "F87171",
        "kpi_text":   "FFB91C1C",
    },
}


def _theme_for(tab_color: str | None) -> dict:
    """Retorna o dicionario de tokens de cor para o tab_color informado.
    Fallback seguro para a paleta verde (primeira da lista)."""
    key = (tab_color or TAB_COLOR).lstrip("#").upper()
    # Tenta match exato, depois match nos 6 primeiros chars (ignora prefixo FF)
    result = TAB_THEMES.get(key) or TAB_THEMES.get(key[-6:])
    if result is None:
        result = TAB_THEMES["047857"]
    return result


HAIRLINE_BOTTOM = Border(bottom=Side(style="thin", color=HAIRLINE))
GROUP_BORDER = Border(
    top=Side(style="thin", color=HAIRLINE),
    bottom=Side(style="thin", color=HAIRLINE),
)
GRAND_BORDER = Border(top=Side(style="double", color=WHITE))
ACCENT_LEFT = Border(left=Side(style="thick", color=NAVY))
CARD_SIDE = Side(style="thin", color=CARD_BORDER)

# Em ordem de prioridade: no modelo, o agrupamento e por "Tipo de Solicitacao"
# mesmo quando tambem existe uma coluna "Situacao".
GROUP_KEYWORDS = ("solicita", "tipo", "categoria")

MONTHS_ORDER = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]

# Layout: banner (1-2), faixa de KPIs (3-4), respiro (5), cabecalho (6), dados (7+)
KPI_VALUE_ROW = 3
KPI_LABEL_ROW = 4
HEADER_ROW = 6
DATA_START_ROW = 7
MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 42
QTY_COL_WIDTH = 12


def _sanitize_tab_name(name: str, used: set[str]) -> str:
    clean = re.sub(r"[\\/?*\[\]:]", "", name or "").strip() or "Planilha"
    clean = clean[:31]
    final_name = clean
    counter = 1
    while final_name.lower() in used:
        suffix = f"_{counter}"
        allowed_len = 31 - len(suffix)
        final_name = f"{clean[:allowed_len]}{suffix}"
        counter += 1
    used.add(final_name.lower())
    return final_name


def _unique_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result = []
    for i, h in enumerate(headers):
        clean = (h or "").strip() or f"Coluna {i + 1}"
        if clean in counts:
            counts[clean] += 1
            clean = f"{clean}_{counts[clean]}"
        else:
            counts[clean] = 0
        result.append(clean)
    return result


def _normalized_header_key(headers: list[str]) -> tuple:
    return tuple(re.sub(r"\s+", " ", (h or "").strip().lower()) for h in headers)


def _merge_same_header_tables(tables: list[dict]) -> list[dict]:
    """Concatena tabelas consecutivas com cabecalhos identicos (relatorios
    multi-pagina repetem o cabecalho a cada pagina)."""
    merged: list[dict] = []
    for table in tables:
        key = _normalized_header_key(table.get("headers") or [])
        if merged and _normalized_header_key(merged[-1]["headers"]) == key:
            merged[-1]["rows"] = merged[-1]["rows"] + (table.get("rows") or [])
        else:
            merged.append(
                {
                    "tableName": table.get("tableName"),
                    "headers": list(table.get("headers") or []),
                    "rows": list(table.get("rows") or []),
                    "count_only": table.get("count_only", False),
                }
            )
    return merged


def _find_group_column(headers: list[str]) -> int | None:
    for kw in GROUP_KEYWORDS:
        for i, h in enumerate(headers):
            if kw in h.lower():
                return i
    return None


def _rows_are_grouped(rows: list, group_col: int) -> bool:
    """True se os valores da coluna de grupo estao contiguos (dados ja vem
    agrupados, como no modelo). Se estiverem intercalados, subtotais por
    grupo virariam ruido e sao omitidos."""
    values = [str(r[group_col]) for r in rows]
    runs = sum(1 for i, v in enumerate(values) if i == 0 or v != values[i - 1])
    return runs == len(set(values))


def _apply_sheet_chrome(ws, *, warn: bool = False, tab_color: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLOR_WARN if warn else (tab_color or TAB_COLOR)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def _fmt_int(n: int) -> str:
    return f"{n:,}".replace(",", ".")


def _compute_period(headers: list[str], parsed_rows: list[list]) -> str | None:
    """Deriva um rotulo de periodo a partir das colunas Ano/Mes, se existirem."""
    ano_col = next((i for i, h in enumerate(headers) if h.strip().lower() == "ano"), None)
    mes_col = next((i for i, h in enumerate(headers) if h.strip().lower() in ("mês", "mes")), None)
    if ano_col is None:
        return None

    years = sorted({int(r[ano_col]) for r in parsed_rows
                    if isinstance(r[ano_col], (int, float)) or str(r[ano_col]).strip().isdigit()})
    if not years:
        return None

    if len(years) == 1 and mes_col is not None:
        months = {str(r[mes_col]).strip().lower() for r in parsed_rows if str(r[mes_col]).strip()}
        present = [m for m in MONTHS_ORDER if m in months]
        if present:
            span = present[0] if len(present) == 1 else f"{present[0]}–{present[-1]}"
            return f"{span} {years[0]}"
        return str(years[0])
    return str(years[0]) if len(years) == 1 else f"{years[0]}–{years[-1]}"


def _build_dynamic_kpis(
    headers: list[str],
    parsed_rows: list[list],
    group_col: int | None,
    n_data_cols: int,
    first_data_row: int,
    last_filter_row: int,
) -> list[tuple[str, str, str | None]]:
    """Constroi KPIs dinamicos com formulas SUBTOTAL/UNIQUE/FILTER ativas contra autofiltro."""
    data_a = f"$A${first_data_row}:$A${last_filter_row}"
    visible = f"SUBTOTAL(3, OFFSET($A${first_data_row}, ROW({data_a})-ROW($A${first_data_row}), 0, 1))"

    # KPI 1: REGISTROS (conta apenas linhas de dados visiveis na coluna A ao filtrar)
    reg_formula = f"=SUBTOTAL(3, A{first_data_row}:A{last_filter_row})"
    kpis: list[tuple[str, str, str | None]] = [(reg_formula, "REGISTROS", "#,##0")]

    # KPI 2: TIPO DE SOLICITAÇÃO (ou grupo - calcula tipos distintos visiveis ao filtrar)
    if group_col is not None:
        gcl = get_column_letter(group_col + 1)
        data_g = f"${gcl}${first_data_row}:${gcl}${last_filter_row}"
        label = re.sub(r"\s*/.*$", "", headers[group_col]).strip().upper()
        label_text = (label[:22] or "TIPOS")
        grp_formula = (
            f'=IFERROR(ROWS(_xlfn.UNIQUE(_xlfn._xlws.FILTER({data_g}, '
            f'({data_a}<>"") * {visible}))), 0)'
        )
        kpis.append((grp_formula, label_text, "#,##0"))
    else:
        kpis.append((_fmt_int(n_data_cols), "COLUNAS", None))

    # KPI 3: PERÍODO (calcula MIN e MAX do Ano entre as linhas visiveis ao filtrar)
    ano_col = next((i for i, h in enumerate(headers) if h.strip().lower() == "ano"), None)
    if ano_col is not None:
        acl = get_column_letter(ano_col + 1)
        ano_range = f"{acl}{first_data_row}:{acl}{last_filter_row}"
        period_formula = (
            f'=IF(SUBTOTAL(5, {ano_range})=0, "", '
            f'IF(SUBTOTAL(5, {ano_range})=SUBTOTAL(4, {ano_range}), '
            f'TEXT(SUBTOTAL(5, {ano_range}), "0"), '
            f'TEXT(SUBTOTAL(5, {ano_range}), "0") & "–" & TEXT(SUBTOTAL(4, {ano_range}), "0")))'
        )
        kpis.append((period_formula, "PERÍODO", None))
    else:
        period = _compute_period(headers, parsed_rows)
        if period:
            kpis.append((period, "PERÍODO", None))

    return kpis


def _write_kpi_strip(
    ws,
    kpis: list[tuple[str, str, str | None]],
    n_cols: int,
    theme: dict,
) -> None:
    """Desenha cartoes de KPI nas linhas 3-4 com formulas dinamicas."""
    kpi_fill_color = theme["kpi_fill"]
    card_bdr_color = theme["card_bdr"]
    accent_color   = theme["accent"]
    kpi_text_color = theme["kpi_text"]

    kpi_accent_left = Border(left=Side(style="thick", color=accent_color))
    card_side = Side(style="thin", color=card_bdr_color)

    ws.row_dimensions[KPI_VALUE_ROW].height = 26
    ws.row_dimensions[KPI_LABEL_ROW].height = 15
    n = len(kpis)
    seg = max(1, (n_cols - (n - 1)) // n)
    spans = []
    for i in range(n):
        c0 = i * (seg + 1) + 1
        c1 = n_cols if i == n - 1 else c0 + seg - 1
        spans.append((c0, min(c1, n_cols)))

    for (value_expr, label, num_fmt), (c0, c1) in zip(kpis, spans):
        if c1 < c0:
            c1 = c0
        for r in (KPI_VALUE_ROW, KPI_LABEL_ROW):
            for c in range(c0, c1 + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=kpi_fill_color)
                cell.border = Border(
                    left=(kpi_accent_left.left if c == c0 else None),
                    top=(card_side if r == KPI_VALUE_ROW else None),
                    bottom=(card_side if r == KPI_LABEL_ROW else None),
                    right=(card_side if c == c1 else None),
                )
        ws.merge_cells(start_row=KPI_VALUE_ROW, start_column=c0, end_row=KPI_VALUE_ROW, end_column=c1)
        ws.merge_cells(start_row=KPI_LABEL_ROW, start_column=c0, end_row=KPI_LABEL_ROW, end_column=c1)
        vcell = ws.cell(row=KPI_VALUE_ROW, column=c0, value=value_expr)
        vcell.font = Font(name=FONT_NAME, size=18, bold=True, color=kpi_text_color)
        vcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if num_fmt:
            vcell.number_format = num_fmt

        lcell = ws.cell(row=KPI_LABEL_ROW, column=c0, value=label)
        lcell.font = Font(name=FONT_NAME, size=8.5, bold=True, color=MUTED)
        lcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _write_dynamic_panel(ws, *, group_col: int, start_col: int, block_start: int,
                         first_data_row: int, last_filter_row: int, unique_count: int,
                         header_short: str, data_font: Font, theme: dict) -> None:
    """Desenha um painel de contabilizacao dinamica (cartao) a partir de `start_col`.
    Colunas: S(+1) = grupo, S+2 = Qtd (com barras de dados), S+3 = %.
    Sempre visivel; recalcula ao filtrar via SUBTOTAL sobre linhas visiveis."""
    banner_color  = theme["banner"]
    kpi_fill_color = theme["kpi_fill"]
    card_bdr_color = theme["card_bdr"]
    bar_hex        = theme["bar_color"]
    kpi_text_color = theme["kpi_text"]
    card_side      = Side(style="thin", color=card_bdr_color)

    type_col = start_col
    qty_col = start_col + 2
    pct_col = start_col + 3
    TL = get_column_letter(type_col)
    QL = get_column_letter(qty_col)
    gcl = get_column_letter(group_col + 1)

    data_a = f"$A${first_data_row}:$A${last_filter_row}"
    data_g = f"${gcl}${first_data_row}:${gcl}${last_filter_row}"
    visible = f"SUBTOTAL(3, OFFSET($A${first_data_row}, ROW({data_a})-ROW($A${first_data_row}), 0, 1))"
    vis_total = f"SUBTOTAL(3, {data_a})"

    # 1. Cabecalho do cartao (cor tematica, contagem ao vivo)
    ws.merge_cells(start_row=block_start, start_column=type_col, end_row=block_start, end_column=pct_col)
    ttl = ws.cell(
        row=block_start, column=type_col,
        value=f'="Contabilização por {header_short}  —  " & TEXT({vis_total}, "#,##0") & " registros"',
    )
    ttl.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    ttl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[block_start].height = 24
    for c in range(type_col, pct_col + 1):
        ws.cell(row=block_start, column=c).fill = PatternFill("solid", fgColor=banner_color)

    # 2. Rotulos de coluna
    label_row = block_start + 1
    label_font = Font(name=FONT_NAME, size=9.5, bold=True, color=kpi_text_color)
    label_border = Border(bottom=Side(style="thin", color=card_bdr_color))
    for c in range(type_col, pct_col + 1):
        lc = ws.cell(row=label_row, column=c)
        lc.fill = PatternFill("solid", fgColor=kpi_fill_color)
        lc.border = label_border
    ws.merge_cells(start_row=label_row, start_column=type_col, end_row=label_row, end_column=type_col + 1)
    hcell = ws.cell(row=label_row, column=type_col, value=header_short)
    hcell.font = label_font
    hcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col, txt in ((qty_col, "Qtd"), (pct_col, "%")):
        cell = ws.cell(row=label_row, column=col, value=txt)
        cell.font = label_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 3. Lista dinamica (spill) + Qtd + %
    spill_row = label_row + 1
    spill_formula = (
        f'=IFERROR(_xlfn.UNIQUE(_xlfn._xlws.FILTER({data_g}, '
        f'({data_a}<>"") * {visible})), "")'
    )
    spill_cell = ws.cell(row=spill_row, column=type_col)
    spill_cell.value = ArrayFormula(f"{TL}{spill_row}", spill_formula)
    spill_cell.font = data_font

    block_end = spill_row + unique_count - 1
    for offset in range(unique_count):
        r = spill_row + offset
        ws.cell(row=r, column=type_col).font = data_font
        qty = ws.cell(
            row=r, column=qty_col,
            value=f'=IF({TL}{r}="", "", SUMPRODUCT(({data_a}<>"") * ({data_g}={TL}{r}) * {visible}))',
        )
        qty.font = data_font
        qty.alignment = Alignment(horizontal="center", vertical="center")
        pct = ws.cell(
            row=r, column=pct_col,
            value=f'=IF(OR({TL}{r}="", {vis_total}=0), "", {QL}{r}/{vis_total})',
        )
        pct.font = Font(name=FONT_NAME, size=10, color=MUTED)
        pct.number_format = "0.0%"
        pct.alignment = Alignment(horizontal="center", vertical="center")

    # 4. Barras de dados na coluna Qtd (cor tematica)
    bar_rule = DataBarRule(start_type="num", start_value=0, end_type="max", color=bar_hex, showValue=True)
    ws.conditional_formatting.add(f"{QL}{spill_row}:{QL}{block_end}", bar_rule)

    # 5. Borda leve de cartao
    for r in range(block_start, block_end + 1):
        left = ws.cell(row=r, column=type_col)
        right = ws.cell(row=r, column=pct_col)
        left.border = Border(left=card_side, top=left.border.top,
                             bottom=(card_side if r == block_end else left.border.bottom))
        right.border = Border(right=card_side, top=right.border.top,
                              bottom=(card_side if r == block_end else right.border.bottom))


def _write_model_sheet(ws, table: dict, sheet_title: str, tab_color: str | None = None) -> int:  # noqa: C901
    """Escreve uma tabela no layout premium. Retorna a contagem de linhas de dados."""
    theme = _theme_for(tab_color)
    banner_color   = theme["banner"]
    banner_dk      = theme["banner_dk"]
    zebra_color    = theme["zebra"]
    group_fill_clr = theme["group_fill"]

    headers = _unique_headers(table.get("headers") or [])
    raw_rows = table.get("rows") or []
    parsed_rows = [
        [cell_value(row[i]) if i < len(row) else "" for i in range(len(headers))]
        for row in raw_rows
    ]

    group_col = _find_group_column(headers)

    # Coluna extra "Quantidade" para as contagens das linhas de total
    qty_header = "Quantidade"
    if any(h.lower() == qty_header.lower() for h in headers):
        qty_header = "Qtd. Registros"
    all_headers = headers + [qty_header]
    n_cols = len(all_headers)
    n_data_cols = len(headers)

    # Colunas numericas (alinhamento, formato e somas nos totais)
    col_is_numeric = []
    for c in range(n_data_cols):
        values = [r[c] for r in parsed_rows if r[c] != ""]
        numeric_values = [v for v in values if isinstance(v, (int, float))]
        col_is_numeric.append(bool(values) and (len(numeric_values) / len(values)) >= 0.5)

    # Relatorio de contagem: os totais apenas contam registros por grupo, sem
    # somar colunas numericas que nao sao medidas (Ano, Dias Atraso, etc.).
    count_only = bool(table.get("count_only"))
    col_should_sum = [False] * n_data_cols if count_only else col_is_numeric

    def num_fmt_for(col_idx: int) -> str:
        header = headers[col_idx].strip().lower()
        if is_currency_column(headers[col_idx]):
            return "R$ #,##0.00"
        # Ano (e afins como codigos/numeros de identificacao) sem separador de
        # milhar: 2026, nao 2.026.
        if header in ("ano", "año") or "ano" in header.split() or header in ("cnpj", "cpf", "codigo", "código"):
            return "0"
        return "#,##0"

    # Larguras por coluna a partir do conteudo; colunas estouradas ganham wrap
    natural_widths = []
    for c in range(n_data_cols):
        lengths = [len(str(headers[c]))] + [len(str(r[c])) for r in parsed_rows if r[c] != ""]
        natural_widths.append((max(lengths) if lengths else 0) + 4)
    wrapped_cols = {c for c, w in enumerate(natural_widths) if w > MAX_COL_WIDTH}
    for c in range(n_data_cols):
        ws.column_dimensions[get_column_letter(c + 1)].width = min(
            max(natural_widths[c], MIN_COL_WIDTH), MAX_COL_WIDTH
        )
    ws.column_dimensions[get_column_letter(n_cols)].width = QTY_COL_WIDTH

    # Oculta a coluna "Responsável" quando solicitado (dados permanecem intactos)
    if table.get("hide_responsavel_col"):
        resp_idx = next((i for i, h in enumerate(headers) if h.strip().lower() == "responsável"), None)
        if resp_idx is not None:
            ws.column_dimensions[get_column_letter(resp_idx + 1)].hidden = True

    _apply_sheet_chrome(ws, tab_color=tab_color)
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"

    # --- Banner tematico (linhas 1-2) ---
    banner_fill = PatternFill("solid", fgColor=banner_color)
    for r in (1, 2):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = banner_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    title = (table.get("tableName") or sheet_title).strip() or sheet_title
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    sub_cell = ws.cell(
        row=2, column=1,
        value=f"Exportado em {exported_at}  |  Total: {len(parsed_rows)} registros",
    )
    sub_cell.font = Font(name=FONT_NAME, size=9.5, color=BANNER_SUBTITLE)
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # (A Faixa de KPIs e desenhada apos calcular a ultima linha da tabela para dinamizar as formulas)
    ws.row_dimensions[HEADER_ROW - 1].height = 6  # respiro antes do cabecalho

    # --- Cabecalho da tabela ---
    header_border = Border(bottom=Side(style="medium", color=banner_dk))
    ws.row_dimensions[HEADER_ROW].height = 26
    header_font = Font(name=FONT_NAME, size=10.5, bold=True, color=WHITE)
    for c, h in enumerate(all_headers):
        is_qty = c == n_data_cols
        if is_qty:
            horizontal = "center"
        else:
            horizontal = "right" if col_is_numeric[c] else "left"
        cell = ws.cell(row=HEADER_ROW, column=c + 1, value=h)
        cell.fill = PatternFill("solid", fgColor=banner_color)
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)

    # --- Dados ---
    data_font = Font(name=FONT_NAME, size=10, color=INK)
    group_font = Font(name=FONT_NAME, size=10, bold=True, color=theme["kpi_text"])
    grand_font = Font(name=FONT_NAME, size=10.5, bold=True, color=WHITE)
    fix_row_height = not wrapped_cols

    def data_alignment(col_idx: int, is_number: bool) -> Alignment:
        return Alignment(
            horizontal="right" if is_number else "left",
            vertical="center",
            wrap_text=col_idx in wrapped_cols,
        )

    # Coluna do rotulo dos totais: primeira coluna de texto a partir da 2a.
    # Nunca a coluna 1 em totais de grupo -- ela precisa ficar vazia para o
    # SUBTOTAL(3, A:A) do TOTAL GERAL nao contar as linhas de total. No TOTAL
    # GERAL (fora do proprio range) a coluna 1 e um fallback seguro.
    _text_col_after_first = next(
        (c + 1 for c in range(1, n_data_cols) if not col_is_numeric[c]), None
    )
    group_label_col = _text_col_after_first or (2 if n_data_cols >= 2 else 1)
    grand_label_col = _text_col_after_first or (
        1 if n_data_cols >= 1 and not col_is_numeric[0] else group_label_col
    )

    grand_fill_clr = banner_color

    def write_total_row(row_idx: int, label: str, group_value, range_start: int, range_end: int, *, grand: bool, group_idx: int = 0):
        gc = GROUP_COLORS[group_idx % len(GROUP_COLORS)]
        fill = PatternFill("solid", fgColor=grand_fill_clr if grand else gc["fill"])
        grp_font = Font(name=FONT_NAME, size=10, bold=True, color=gc["font"]) if not grand else grand_font
        border = GRAND_BORDER if grand else GROUP_BORDER
        for c in range(n_cols):
            cell = ws.cell(row=row_idx, column=c + 1)
            cell.fill = fill
            cell.font = grp_font
            cell.border = border
            if c == n_cols - 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c < n_data_cols and col_is_numeric[c]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        label_col = grand_label_col if grand else group_label_col
        ws.cell(row=row_idx, column=label_col, value=label)
        if group_value is not None and group_col is not None:
            if group_col + 1 == label_col:
                ws.cell(row=row_idx, column=label_col, value=f"{label} — {group_value}")
            else:
                ws.cell(row=row_idx, column=group_col + 1, value=group_value)
        ws.cell(row=row_idx, column=n_cols, value=f"=SUBTOTAL(3, A{range_start}:A{range_end})")
        for c in range(n_data_cols):
            if col_should_sum[c] and c + 1 != label_col:
                col_letter = get_column_letter(c + 1)
                cell = ws.cell(
                    row=row_idx, column=c + 1,
                    value=f"=SUBTOTAL(109, {col_letter}{range_start}:{col_letter}{range_end})",
                )
                cell.number_format = num_fmt_for(c)
        if grand:
            ws.row_dimensions[row_idx].height = 22

    subtotals_on = group_col is not None and bool(parsed_rows) and _rows_are_grouped(parsed_rows, group_col)

    current_row = DATA_START_ROW
    group_start_row = DATA_START_ROW
    prev_group_value = None
    group_idx = 0          # indice rotativo para GROUP_COLORS
    row_in_group = 0       # contador de linhas dentro do grupo atual

    for i, row_vals in enumerate(parsed_rows):
        group_value = str(row_vals[group_col]) if group_col is not None else None

        if subtotals_on and i > 0 and group_value != prev_group_value:
            write_total_row(current_row, "Total", prev_group_value, group_start_row, current_row - 1, grand=False, group_idx=group_idx)
            current_row += 1
            group_start_row = current_row
            group_idx += 1
            row_in_group = 0

        if fix_row_height:
            ws.row_dimensions[current_row].height = 19

        # Zebra colorida por grupo: cada grupo usa seu tom de zebra
        gc = GROUP_COLORS[group_idx % len(GROUP_COLORS)] if subtotals_on else GROUP_COLORS[0]
        striped = (row_in_group % 2 == 1)
        group_accent_side = Side(style="thin", color=gc["accent"])

        for c, val in enumerate(row_vals):
            cell = ws.cell(row=current_row, column=c + 1, value=val)
            cell.font = data_font
            if striped:
                cell.fill = PatternFill("solid", fgColor=gc["zebra"])
            # Borda lateral colorida na primeira coluna para identidade do grupo
            if c == 0 and subtotals_on:
                cell.border = Border(left=group_accent_side, bottom=Side(style="thin", color=HAIRLINE))
            if isinstance(val, (int, float)):
                cell.alignment = data_alignment(c, True)
                cell.number_format = num_fmt_for(c)
            else:
                cell.alignment = data_alignment(c, False)

        prev_group_value = group_value
        row_in_group += 1
        current_row += 1

    if subtotals_on:
        write_total_row(current_row, "Total", prev_group_value, group_start_row, current_row - 1, grand=False, group_idx=group_idx)
        current_row += 1

    last_filter_row = current_row - 1  # autofiltro cobre dados + totais de grupo

    grand_total_row = current_row
    write_total_row(grand_total_row, "TOTAL GERAL", None, DATA_START_ROW, last_filter_row, grand=True)
    current_row += 1

    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{last_filter_row}"
    ws.freeze_panes = f"A{DATA_START_ROW}"

    # --- Faixa de KPIs Dinamicos (linhas 3-4) ---
    # Agora calculada com formulas ativas contra o autofiltro (SUBTOTAL/UNIQUE/FILTER)
    dynamic_kpis = _build_dynamic_kpis(
        headers, parsed_rows, group_col, n_data_cols, DATA_START_ROW, last_filter_row
    )
    _write_kpi_strip(ws, dynamic_kpis, n_cols, theme)

    # --- Paineis de contabilizacao dinamica (cartoes de dashboard) ---
    # Sempre visiveis: mostram a distribuicao completa e recalculam ao filtrar.
    if group_col is not None and parsed_rows:
        block_start = grand_total_row + 3
        _write_dynamic_panel(
            ws, group_col=group_col, start_col=1, block_start=block_start,
            first_data_row=DATA_START_ROW, last_filter_row=last_filter_row,
            unique_count=len({str(r[group_col]) for r in parsed_rows}),
            header_short=re.sub(r"\s*/.*$", "", headers[group_col]).strip(),
            data_font=data_font, theme=theme,
        )

        # Segundo painel por "Situacao", lado a lado (colunas F:I), se existir
        # uma coluna de situacao distinta da coluna de grupo e houver espaco.
        situacao_col = next(
            (i for i, h in enumerate(headers) if "situa" in h.lower() and i != group_col), None
        )
        if situacao_col is not None and n_cols >= 9:
            _write_dynamic_panel(
                ws, group_col=situacao_col, start_col=6, block_start=block_start,
                first_data_row=DATA_START_ROW, last_filter_row=last_filter_row,
                unique_count=len({str(r[situacao_col]) for r in parsed_rows}),
                header_short=re.sub(r"\s*/.*$", "", headers[situacao_col]).strip(),
                data_font=data_font, theme=theme,
            )

    return len(parsed_rows)


def _write_warning_sheet(ws, message: str) -> None:
    _apply_sheet_chrome(ws, warn=True)
    ws.column_dimensions["A"].width = 90
    title = ws.cell(row=1, column=1, value="⚠  Não foi possível extrair tabelas deste PDF")
    title.font = Font(name=FONT_NAME, size=12, bold=True, color=INK)
    ws.row_dimensions[1].height = 24
    reason = ws.cell(row=2, column=1, value=message)
    reason.font = Font(name=FONT_NAME, size=10, color=MUTED)
    reason.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _proc_key(val) -> str:
    """Chave normalizada de um numero de processo para cruzamento entre abas."""
    if val in (None, ""):
        return ""
    return re.sub(r"\s+", "", str(val)).lower()


def _inject_responsavel(main_table: dict, proc_to_persons: dict[str, list[str]]) -> None:
    """Adiciona a coluna 'Responsavel' na tabela principal: para cada processo,
    junta os nomes das abas (pessoas) que tambem contem aquele processo."""
    headers = main_table.get("headers") or []
    if "responsável" in [h.strip().lower() for h in headers]:
        return
    main_table["headers"] = list(headers) + ["Responsável"]
    main_table["rows"] = [
        list(row) + [", ".join(proc_to_persons.get(_proc_key(row[0]) if row else "", []))]
        for row in main_table.get("rows", [])
    ]


def build_workbook(
    file_results: list[dict],
    hide_responsavel: bool = False,
    hide_secondary_tabs: bool = False,
) -> tuple[bytes, list[dict]]:
    """file_results: lista de {fileName, tabName, tables, extractionMethod, warning}
    onde `tables` e uma lista de {tableName, headers, rows}. O primeiro arquivo e
    o principal. `hide_responsavel` oculta a coluna "Responsável" na aba principal;
    `hide_secondary_tabs` oculta as abas dos arquivos secundarios.
    Retorna (bytes_do_xlsx, lista_de_resumo_por_arquivo)."""
    wb = Workbook()
    wb.remove(wb.active)

    used_tab_names: set[str] = set()

    # Passagem 1: mescla tabelas e reserva os nomes de aba na ordem, para que
    # a aba principal conheca os nomes das abas de responsaveis antes de montar.
    prepared = []
    for result in file_results:
        base_name = result.get("tabName") or result["fileName"]
        tables = _merge_same_header_tables(
            [t for t in (result.get("tables") or []) if t.get("headers")]
        )
        if not tables:
            names = [_sanitize_tab_name(base_name, used_tab_names)]
            ordered = []
        else:
            main_idx = max(range(len(tables)), key=lambda i: len(tables[i].get("rows") or []))
            ordered = [tables[main_idx]] + [t for i, t in enumerate(tables) if i != main_idx]
            names = [
                _sanitize_tab_name(base_name if i == 0 else f"{base_name} ({i + 1})", used_tab_names)
                for i in range(len(ordered))
            ]
        prepared.append({"result": result, "tables": tables, "ordered": ordered, "names": names})

    # Mapa processo -> responsaveis (abas secundarias que contem o processo)
    proc_to_persons: dict[str, list[str]] = {}
    for p in prepared[1:]:
        if not p["ordered"]:
            continue
        person = p["names"][0]
        for key in {_proc_key(row[0]) for row in p["ordered"][0].get("rows", []) if row and _proc_key(row[0])}:
            proc_to_persons.setdefault(key, []).append(person)

    # Injeta "Responsavel" na tabela principal, se houver responsaveis com dados
    if proc_to_persons and prepared and prepared[0]["ordered"]:
        main_table = prepared[0]["ordered"][0]
        _inject_responsavel(main_table, proc_to_persons)
        if hide_responsavel:
            main_table["hide_responsavel_col"] = True

    # Passagem 2: monta as abas
    summary = []
    tab_index = 0
    for file_idx, p in enumerate(prepared):
        result, names, ordered = p["result"], p["names"], p["ordered"]
        # Arquivos secundarios (indice > 0) podem ter suas abas ocultadas.
        is_secondary = file_idx > 0
        sheets = []
        total_data_rows = 0
        if not ordered:
            ws = wb.create_sheet(title=names[0])
            _write_warning_sheet(ws, result.get("warning") or "Nenhuma tabela identificada neste PDF.")
            sheets.append(ws)
        else:
            for name, table in zip(names, ordered):
                ws = wb.create_sheet(title=name)
                color = TAB_PALETTE[tab_index % len(TAB_PALETTE)]
                tab_index += 1
                total_data_rows += _write_model_sheet(ws, table, name, tab_color=color)
                sheets.append(ws)

        if is_secondary and hide_secondary_tabs:
            for ws in sheets:
                ws.sheet_state = "hidden"

        summary.append(
            {
                "fileName": result["fileName"],
                "tabName": names[0],
                "tableCount": len(p["tables"]),
                "rowCount": total_data_rows,
                "extractionMethod": result.get("extractionMethod", "digital"),
                "warning": result.get("warning"),
            }
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    # Converte os arrays CSE legados (UNIQUE/FILTER) em arrays dinamicos, senao
    # o Excel os confina a uma celula e a lista de tipos nao derrama.
    xlsx_bytes = inject_dynamic_arrays(buffer.getvalue())
    return xlsx_bytes, summary
